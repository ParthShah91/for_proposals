from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Create a new Excel workbook and add a worksheet
workbook = Workbook()
worksheet = workbook.active
worksheet.title = 'MySheet'

# Define some cell styles for formatting
bold_font = Font(bold=True)
italic_font = Font(italic=True)
underline_font = Font(underline='single')
red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
green_fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
blue_fill = PatternFill(start_color='FF0000FF', end_color='FF0000FF', fill_type='solid')
center_alignment = Alignment(horizontal='center')
right_alignment = Alignment(horizontal='right')
font_color = Font(color="FFFF0000")  # Red font color

# Write data to cells with various formatting, colors, and alignment
worksheet.cell(row=1, column=1, value='Regular Text')
worksheet.cell(row=2, column=1, value='Bold Text').font = bold_font
worksheet.cell(row=3, column=1, value='Italic Text').font = italic_font
worksheet.cell(row=4, column=1, value='Underline Text').font = underline_font
worksheet.cell(row=5, column=1, value='Red Background').fill = red_fill
worksheet.cell(row=6, column=1, value='Green Background').fill = green_fill
worksheet.cell(row=7, column=1, value='Blue Background').fill = blue_fill
worksheet.cell(row=8, column=1, value='Centered Text').alignment = center_alignment
worksheet.cell(row=9, column=1, value='Right-aligned Text').alignment = right_alignment
worksheet.cell(row=10, column=1, value='Red Font Color').font = font_color

# Adding real data alongside formulas in some cells
worksheet.cell(row=1, column=2, value=42)  # Real data
worksheet.cell(row=2, column=2, value=7)  # Real data
worksheet.cell(row=3, column=2, value=3.14)  # Real data
worksheet.cell(row=4, column=2, value='Hello')  # Real data
worksheet.cell(row=5, column=2, value=100)  # Real data
worksheet.cell(row=6, column=2, value=18)  # Real data
worksheet.cell(row=7, column=2, value='World')  # Real data
worksheet.cell(row=8, column=2, value=123.45)  # Real data
worksheet.cell(row=9, column=2, value=99)  # Real data

# Adding formulas in some cells
worksheet.cell(row=11, column=1, value='Formula Sum')
worksheet.cell(row=11, column=2, value='=SUM(B1:B10)')

# Save the workbook to a file
workbook.save('formatted_data_with_real_data.xlsx')

print('Data written to formatted_data_with_real_data.xlsx')
